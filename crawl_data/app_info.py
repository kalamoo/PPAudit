from openpyxl import load_workbook
import re, string
from pathlib import Path
import hashlib


name_dict = {
        "psvr":5,
        "viveport":5,
        "microsoftvr1":5,
        "microsoftvr2":3,
        "quest-app":5,
        "quest-game":5,
        "quest-entertainment":5,
        "rift":5,
        "go":5,
        "gear":5,
        "steamvr":5,
        "sidequest":5,
    }

platform_dict = {
        "psvr": "PSVR",
        "viveport": "Viveport",
        "microsoftvr1": "Microsoft",
        "microsoftvr2": "Microsoft",
        "quest-app": "Quest",
        "quest-game": "Quest",
        "quest-entertainment": "Quest",
        "rift": "Rift",
        "go": "Go",
        "gear": "Gear",
        "steamvr": "SteamVR",
        "sidequest": "Sidequest",
    }

price_dict = {
    "psvr":8,  # Free / $9.99 / Not available for purchase
    "viveport":18,  # Free / Play / Y9.99 / Null
    "microsoftvr1":9,  # Free / $9.99
    "microsoftvr2":0,  # no price
    "quest-app":7,  # Get(means Free) / $9.99 / Purchased / Null
    "quest-game":7,
    "quest-entertainment":7,
    "rift":7,
    "go":7,
    "gear":7,
    "steamvr":3,  # F(f)ree T(t)o P(p)lay / $9.99 / Free / Null
    "sidequest":3,  # PAID / FREE
}

rating_dict = {  # 
    "psvr":0,  # no rating
    "viveport":16,  # 5 / null / Null
    "microsoftvr1":0,  # no rating
    "microsoftvr2":0,  # no rating
    "quest-app":0,  # no rating
    "quest-game":0,
    "quest-entertainment":0,
    "rift":0,
    "go":0,
    "gear":0,
    "steamvr":0,
    "sidequest":3,  #  X.XX
}

num_rating_dict = {
    "psvr":0,  # num rating
    "viveport":17,  # 5 review(s) / null
    "microsoftvr1":10,  # '' / 8
    "microsoftvr2":0,  # no num rating
    "quest-app":8,  # '' / XX ratings
    "quest-game":8,
    "quest-entertainment":8,
    "rift":8,
    "go":8,
    "gear":8,
    "steamvr":13,  # (XX reviews)
    "sidequest":12,  # null / XXk
}


item_page_link_dict = {
        "psvr": 4,
        "viveport": 4,
        "microsoftvr1": 4,
        "microsoftvr2": 4,
        "quest-app": 4,
        "quest-game": 4,
        "quest-entertainment": 4,
        "rift": 4,
        "go": 4,
        "gear": 4,
        "steamvr": 4,
        "sidequest": 4,
    }

website_link_dict = {
        "psvr": 0,  # no website link
        "viveport": 0,  # no website link
        "microsoftvr1": 0,  # no website link
        "microsoftvr2": 16, 
        "quest-app": 10,
        "quest-game": 10,
        "quest-entertainment": 10,
        "rift": 10,
        "go": 10,
        "gear": 11,
        "steamvr": 14,
        "sidequest": 15,  # app-lab=Yes -> website link showed on its oculus item page 
    }

pp_link_dict = {
        "psvr": 11,
        "viveport": 8,
        "microsoftvr1": 8,
        "microsoftvr2": 12,
        "quest-app": 13,
        "quest-game": 13,
        "quest-entertainment": 13,
        "rift": 13,
        "go": 13,
        "gear": 14,
        "steamvr": 15,
        "sidequest": 18,  # app-lab=Yes -> website link showed on its oculus item page 
    }

publisher_dict = {
        "psvr": 6,
        "viveport": 20,
        "microsoftvr1": 6,
        "microsoftvr2": 10,
        "quest-app": 14,
        "quest-game": 14,
        "quest-entertainment": 14,
        "rift": 14,
        "go": 14,
        "gear": 10,
        "steamvr": 9,
        "sidequest": 0,  # no publisher
    }

generes_dict = {
        "psvr": 7,
        "viveport": 6,
        "microsoftvr1": 13,
        "microsoftvr2": 8,
        "quest-app": 14,  # within app info dict 
        "quest-game": 14,
        "quest-entertainment": 14, 
        "rift": 14, 
        "go": 14,
        "gear": 10,
        "steamvr": 7,  # tags
        "sidequest": 11,  # tags
    }

description_dict = {
        "psvr": 9,
        "viveport": 12,
        "microsoftvr1": 12,
        "microsoftvr2": 6,
        "quest-app": 6,
        "quest-game": 6,
        "quest-entertainment": 6,
        "rift": 6,
        "go": 6,
        "gear": 6,
        "steamvr": 6,
        "sidequest": 6,
    }


class MetaInfoReader:
# NOTE: I really shoudn't export the raw meta-info in xlsx format,
# which complicates the merge process a little bit
    def __init__(self, source, file_path):
        self.source = source
        self.file_path = file_path
        self.sheetAll = load_workbook(file_path)
        self.sheet = self.sheetAll.worksheets[0]
        self.max_row = self.sheet.max_row
    
    def text_normalize(self, text, mode):
        if not text:
            return None
        if mode == 'raw':
            return text
        if mode == 'complex':
            # # punctuations !"#$%&'()*+,-./:;<=>?@[\]^_`{|}~
            p = re.compile('[^a-zA-Z0-9\s' + re.escape(string.punctuation) + ']')
            text = re.sub(p, '', text)  # 1. ignore character that is not in legal character
            text = re.sub(r'(\s+)', ' ', text)  # merge multiple \s to singe space
            return text
        if mode == 'simple':
            text = text.lower()
            text = re.sub(r'[^a-z0-9 ]+', '', text)  # only save digit + a-z + space
            text = re.sub(r'(\s+)', ' ', text)  # merge spaces
            return text
    
    def link_check(self, link):
        if not link:
            return None
        link = link.lower()
        if not (link.startswith('http') or (re.sub(r'[a-z0-9-.]', '', link)=='' and '.' in link)):
            return None
        return link

    def get_name(self, item_idx, mode):
        # mode： 'raw', 'complex', 'simple'
        name_raw = self.sheet.cell(item_idx, name_dict[self.source]).value
        return self.text_normalize(name_raw, mode)
    
    def get_platform(self, item_idx):
        if self.source == "sidequest" and self.sheet.cell(item_idx, 16).value == "Yes":
            return "App Lab"
        return platform_dict[self.source]

    def get_free_or_paid(self, item_idx):
        if self.source == 'psvr':
            if self.sheet.cell(item_idx, price_dict[self.source]).value.lower() == 'free':
                return "Free"
            else:
                return "Paid"

        if self.source == "viveport":
            if self.sheet.cell(item_idx, price_dict[self.source]).value.lower() in ['free', 'play']:
                return "Free"
            else:
                return "Paid"

        if self.source == "microsoftvr1":
            if self.sheet.cell(item_idx, price_dict[self.source]).value.lower() == 'free':
                return "Free"
            else:
                return "Paid"
        if self.source == "microsoftvr2":
            return "Paid"

        if self.get_platform(item_idx) in ["Quest", "Rift", "Go", "Gear"]:
            if self.sheet.cell(item_idx, price_dict[self.source]).value.lower() in ['get', 'purchased']:
                return "Free"
            else:
                return "Paid"

        if self.source == "steamvr":
            if 'free' in self.sheet.cell(item_idx, price_dict[self.source]).value.lower():
                return "Free"
            else:
                return "Paid"

        if self.source == "sidequest":
            if 'FREE' in self.sheet.cell(item_idx, price_dict[self.source]).value:
                return "Free"
            else:
                return "Paid"
    
    def get_rating(self, item_idx):
        if rating_dict[self.source] == 0:
            return None

        if self.source == "viveport":
            if self.sheet.cell(item_idx, rating_dict[self.source]).value.lower() in ['free', 'play', '', 'null', '']:
                return None
            else:
                return float(self.sheet.cell(item_idx, rating_dict[self.source]).value.lower())

        if self.source == "sidequest":
            pattern = r'\d{1}\.\d{2}'  # match X.XX ratings
            rating = re.findall(pattern, self.sheet.cell(item_idx, rating_dict[self.source]).value)
            if len(rating) > 0:
                return float(rating[0])
            else:
                return None

    def get_num_rating(self, item_idx):
        if num_rating_dict[self.source] == 0:
            return None
    
        if self.source == "viveport":
            num_str = self.sheet.cell(item_idx, num_rating_dict[self.source]).value.lower().split(' ')[0]
            if num_str in ['', 'null']:
                return None
            else:
                return int(num_str)
                
        if self.source == "microsoftvr1":
            num_str = self.sheet.cell(item_idx, num_rating_dict[self.source]).value.lower() 
            if num_str in ['', 'null']:
                return None
            else:
                if num_str[-1] == 'k':
                    num_str = num_str.replace('k', '')
                    num_ = int(float(num_str) * 1000)
                elif num_str[-1] == 'm':
                    num_str = num_str.replace('m', '')
                    num_ = int(float(num_str) * 1000000)
                else:
                    num_ = int(num_str)
                return num_
    
        if self.get_platform(item_idx) in ["Quest", "Rift", "Go", "Gear"]:
            if self.sheet.cell(item_idx, num_rating_dict[self.source]).value:
                return int(self.sheet.cell(item_idx, num_rating_dict[self.source]).value.lower().split(' ')[0].replace(',', ''))
            else:
                return None
            
        if self.source == "steamvr":
            if self.sheet.cell(item_idx, num_rating_dict[self.source]).value:
                pattern = r'\((.*?)\)'  # match content within '()'
                matches = re.findall(pattern, self.sheet.cell(item_idx, num_rating_dict[self.source]).value.lower())
                if len(matches) > 0:
                    return int(matches[0].split(' ')[0].replace(',', ''))
            return None
    
        if self.source == "sidequest":
            if self.sheet.cell(item_idx, num_rating_dict[self.source]).value.lower() in ['', 'null']:
                return None
            else:
                num_str = self.sheet.cell(item_idx, num_rating_dict[self.source]).value.lower()
                if num_str[-1] == 'k':
                    num_str = num_str.replace('k', '')
                    num_ = int(float(num_str) * 1000)
                elif num_str[-1] == 'm':
                    num_str = num_str.replace('m', '')
                    num_ = int(float(num_str) * 1000000)
                else:
                    num_ = int(num_str)
                return num_
    
    def get_item_page_link(self, item_idx):
        # print(self.sheet.cell(item_idx, item_page_link_dict[self.source]).value)
        link = self.sheet.cell(item_idx, item_page_link_dict[self.source]).value
        return self.link_check(link)

    def get_website_link(self, item_idx):
        if website_link_dict[self.source] == 0:
            return None
        if self.get_platform(item_idx) == "App Lab":
            link_ = self.sheet.cell(item_idx, 17).value
            return self.link_check(link_)
        link_ = self.sheet.cell(item_idx, website_link_dict[self.source]).value
        return self.link_check(link_)

    def get_pp_link(self, item_idx):
        if self.source == "viveport":
            pp_link = self.sheet.cell(item_idx, 8).value
            terms_link = self.sheet.cell(item_idx, 10).value
            if pp_link and pp_link != "null":
                return self.link_check(pp_link)
            if terms_link and terms_link != "https://www.viveport.com/terms-of-use": 
                return self.link_check(terms_link)
            return None

        if self.source == "microsoftvr2":
            pp_link = self.sheet.cell(item_idx, 12).value
            terms_link = self.sheet.cell(item_idx, 14).value
            if pp_link and pp_link != "https://privacy.microsoft.com/en-us":
                return self.link_check(pp_link)
            if terms_link and terms_link != "https://www.microsoft.com/en-US/store/b/terms-of-sale":
                return self.link_check(terms_link)
            return None

        if self.source == "sidequest":
            if self.get_platform(item_idx) == "App Lab":
                return self.link_check(self.sheet.cell(item_idx, 18).value)
            else:
                return self.link_check(self.sheet.cell(item_idx, 19).value)

        pp_link = self.sheet.cell(item_idx, pp_link_dict[self.source]).value

        return self.link_check(pp_link)

    def get_publisher_quest(self, item_idx):
        info_details = self.sheet.cell(item_idx, publisher_dict[self.source]).value
        if self.source == "gear":
            info_details = re.sub(r"app infos", "app info", info_details)
        info_details_list = eval(info_details)
        for info_detail in info_details_list:  # {"app info":"PublisherSurvios"}
            if info_detail['app info'].find("Publisher") != -1:
                publisher = info_detail['app info'].split("Publisher")[-1]
                return self.text_normalize(publisher)
            if info_detail['app info'].find("Developer") != -1:
                developer = info_detail['app info'].split("Developer")[-1]
                return developer
        return None

    def get_publisher(self, item_idx, mode):
        if publisher_dict[self.source] == 0:
            return None

        if self.source == "viveport":
            publisher = self.sheet.cell(item_idx, 20).value
            developer = self.sheet.cell(item_idx, 19).value
            if publisher and publisher != "null":
                return self.text_normalize(publisher, mode)
            return self.text_normalize(developer, mode)

        if self.source == "microsoftvr2":
            info_details = self.sheet.cell(item_idx, publisher_dict[self.source]).value
            info_details = re.sub(r"(\\n*)", " ", info_details)
            info_details = re.sub(r"(\s+)", " ", info_details)
            # print("[DEBUG] \t {}".format(info_details))
            info_details_list = eval(info_details)
            for info_detail in info_details_list:
                if info_detail['info'].find("Published by") != -1:
                    publisher = ' '.join(info_detail['info'].split(' ')[2:-1])
                    return self.text_normalize(publisher, mode)
                if info_detail['info'].find("Developed by") != -1:
                    developer = ' '.join(info_detail['info'].split(' ')[2:-1])
                    return self.text_normalize(developer, mode)
            return None

        if self.get_platform(item_idx) in ["Rift", "Go", "Gear", "Quest"]:
            return self.text_normalize(self.get_publisher_quest(item_idx), mode)

        if self.source == "steamvr":
            info_details = self.sheet.cell(item_idx, publisher_dict[self.source]).value
            if info_details is None:
                return None
            info_details = re.sub(r"\s{2,}", "\n", info_details)
            info_details_list = info_details.split('\n')
            publisher, developer = None, None
            for i in range(len(info_details_list)):
                if info_details_list[i] == "Publisher:":
                    publisher = info_details_list[i+1]
                if info_details_list[i] == "Developer:":
                    developer = info_details_list[i+1]
            if publisher:
                return self.text_normalize(publisher, mode)
            if developer:
                return self.text_normalize(developer, mode)
            return None

        return self.text_normalize(self.sheet.cell(item_idx, publisher_dict[self.source]).value, mode)
    
    def get_genres(self, item_idx):  # list
        genres_string = self.get_genres_split_by_comma(item_idx)
        genres_list = genres_string.split(', ')
        clean_genres_list = []
        for genre in genres_list:
            genre_norm = self.text_normalize(genre, 'simple')
            if not genre_norm or genre_norm == '':
                continue
            clean_genres_list.append(genre_norm)
        genres = list(set(clean_genres_list))
        if len(genres) == 0:
            return None
        return genres
        
    def get_genres_quest(self, item_idx):
        info_details = self.sheet.cell(item_idx, generes_dict[self.source]).value
        if self.source == "gear":
            info_details = re.sub("app infos", "app info", info_details)
        info_details_list = eval(info_details)
        for info_detail in info_details_list:  #  {"app infos":"GenresAdventure, Educational, Exploration"}
            if info_detail['app info'].find("Genre") != -1:
                genres = info_detail['app info'].split("Genres")[-1]
                return genres
        return None
    
    def get_genres_split_by_comma(self, item_idx):
        if self.source == "viveport":
            tags_raw = self.sheet.cell(item_idx, generes_dict[self.source]).value
            tags1 = re.sub(r"\s{2,}", "\n", tags_raw)
            tags = ", ".join(tags1.split("\n"))
            return tags

        if self.get_platform(item_idx) in ["Rift", "Go", "Gear", "Quest"]:
            return self.get_genres_quest(item_idx)

        if self.source == "steamvr":
            # steamvr genres = genres + tags
            # get genres from info dict
            info_details = self.sheet.cell(item_idx, 9).value
            if info_details is None:
                return ''
            info_details = re.sub(r"\s{2,}", "\n", info_details)
            info_details_list = info_details.split('\n')
            genres = ""
            for info_detail in info_details_list:
                if info_detail.find("Genre") != -1:
                    genres = info_detail.split("Genre: ")[-1]

            # get tags
            tags_raw = self.sheet.cell(item_idx, generes_dict[self.source]).value
            tags1 = re.sub(r"\s{2,}", "\n", tags_raw)
            tags = ", ".join(tags1.split("\n")[:-1])

            return genres + ", " + tags

        if self.source == "sidequest":
            tags_raw = self.sheet.cell(item_idx, generes_dict[self.source]).value
            tag_items_list = eval(tags_raw)
            tags = []
            for tage_item in tag_items_list:
                tags.append(tage_item['tages'])
            return ", ".join(tags)

        return self.sheet.cell(item_idx, generes_dict[self.source]).value

    def get_description(self, item_idx, mode): # complex text normalize
        description = self.sheet.cell(item_idx, description_dict[self.source]).value
        if self.source == "sidequest":
            description = description[11:]
        return self.text_normalize(description, mode)

    def get_info_dict(self, item_idx):
        info_dict = {
            'name_simple': self.get_name(item_idx, 'simple'),
            'name_complex': self.get_name(item_idx, 'complex'),
            'name_raw': self.get_name(item_idx, 'raw'),
            'publisher_simple': self.get_publisher(item_idx, 'simple'),
            'publisher_complex': self.get_publisher(item_idx, 'complex'),
            'publisher_raw': self.get_publisher(item_idx, 'raw'),
            'platform': self.get_platform(item_idx),
            'app_link_on_platform': self.get_item_page_link(item_idx),
            'app_link_home': self.get_website_link(item_idx),
            'pp_link': self.get_pp_link(item_idx),
            'genres': self.get_genres(item_idx),
            'free_or_paid': self.get_free_or_paid(item_idx),
            'rating': self.get_rating(item_idx),
            'num_rating': self.get_num_rating(item_idx),
            'description_simple': self.get_description(item_idx, 'simple'),
            'description_complex': self.get_description(item_idx, 'complex'),
            'description_raw': self.get_description(item_idx, 'raw'),
        }

        # make sure platform is one of the following
        assert info_dict['platform'] in ["PSVR", "Viveport", "Microsoft", "Quest", "Rift", "Go",  "Gear", "SteamVR", "Sidequest", "App Lab"]
        
        # make sure link is:
        # 1. None OR
        # 2. str startswith http OR
        # 3. str comply with domain name format（a-z + digits + '.' + '-'; must contan '.')
        for link_type in ['app_link_on_platform', 'app_link_home', 'pp_link']:
            link = info_dict[link_type]
            assert (not link or link.startswith('http') or (re.sub(r'[a-z0-9-.]', '', link)=='' and '.' in link))

        # make sure genres is
        # 1. None OR
        # 2. non-empty list
        assert (not info_dict['genres'] or (isinstance(info_dict['genres'], list) and len(info_dict['genres']) > 0))

        # make sure free or paid is one of the following
        assert info_dict['free_or_paid'] in ['Paid', 'Free']

        # make sure rating is
        # 1. None OR
        # 2. float
        assert (not info_dict['rating'] or isinstance(info_dict['rating'], float))

        # make sure num rating is
        # 1. None OR 
        # 2. int
        assert (not info_dict['num_rating'] or isinstance(info_dict['num_rating'], int))

        return info_dict
    

structured_json = dict()
for source, xlsx_file in source_xlsx_dict.items():
    excel_reader = MetaInfoReader(source=source, file_path=xlsx_file)
    for read_idx in range(2, excel_reader.max_row):
        info_dict = excel_reader.get_info_dict(read_idx)
        # only save those with 'app_link_on_platform'
        if info_dict['app_link_on_platform']:
            sha_id = hashlib.sha256(info_dict['app_link_on_platform'].encode('utf-8')).hexdigest()
            structured_json[sha_id] = info_dict